"""
Génère un fichier Excel de méthodologie de bench IA — BoxIA vs ChatGPT vs Gemini.

Usage : python3 gen-bench-xlsx.py [output.xlsx]
Default output : ./methodologie-bench-IA.xlsx

Inclut tests texte + multimodaux (vision, génération image, voix, conversation).
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else "methodologie-bench-IA.xlsx"

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ---- Styles ----
HEADER = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
THIN = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
LEVEL_COLORS = {
    "Easy":   PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Medium": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Hard":   PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def make_prompts_sheet(name, emoji, prompts):
    ws = wb.create_sheet(f"{emoji} {name}")
    headers = [
        "ID", "Niveau", "Prompt", "Préparation\n(multimodal / RAG)",
        "Réponse attendue / Mots-clés",
        "Note BoxIA\n/5", "Note ChatGPT\n/5", "Note Gemini\n/5",
        "Commentaire",
    ]
    widths = [6, 10, 60, 28, 45, 11, 12, 12, 40]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    for r, p in enumerate(prompts, 2):
        cells = [
            p["id"], p["level"], p["prompt"],
            p.get("multimodal", ""), p.get("expected", ""),
            "", "", "",
            "",
        ]
        for col, val in enumerate(cells, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.alignment = WRAP
            c.border = THIN
            if col == 2:
                c.fill = LEVEL_COLORS.get(p["level"], PatternFill())
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 70
    return ws


# ============================================================================
# ONGLET 1 : MÉTHODOLOGIE
# ============================================================================
ws = wb.create_sheet("📘 Méthodologie")
ws["A1"] = "Méthodologie de bench IA — BoxIA vs ChatGPT vs Gemini"
ws["A1"].font = Font(bold=True, size=16)
ws.merge_cells("A1:A1")
ws.column_dimensions["A"].width = 110

intro = [
    "",
    "Ce fichier vous guide pour comparer rigoureusement les 3 IA sur des cas usage TPE/PME concrets.",
    "",
    "🎯 Objectif",
    "Décider en connaissance de cause sur quels cas vous gardez votre BoxIA en local et quand vous",
    "recommandez à vos collaborateurs d'utiliser ChatGPT ou Gemini en complément.",
    "",
    "📋 Comment utiliser ce fichier",
    "1. Pour chaque onglet, lisez la colonne PROMPT et copiez-le dans les 3 IA (BoxIA, ChatGPT, Gemini).",
    "2. Si le prompt est multimodal (image, voix, génération d'image…), suivez la colonne PRÉPARATION.",
    "3. Notez chaque réponse de 0 à 5 (légende ci-dessous) dans les 3 colonnes NOTE.",
    "4. Ajoutez un commentaire libre dans la dernière colonne.",
    "5. À la fin, l'onglet 📊 SCORE FINAL agrège le tableau.",
    "",
    "🏆 Légende des notes (0-5)",
    "  0 = Réponse fausse, dangereuse ou hors-sujet",
    "  1 = Très partielle, beaucoup d'erreurs",
    "  2 = Acceptable mais à corriger sérieusement avant usage",
    "  3 = Correcte, utilisable telle quelle pour usage interne",
    "  4 = Très bonne, prête pour usage client/officiel",
    "  5 = Excellente, va au-delà de la demande (analyse, contexte, conseils stratégiques)",
    "",
    "🚦 Niveau de difficulté du prompt",
    "  Easy   = 1 étape, vocabulaire courant",
    "  Medium = 2-4 étapes, vocabulaire métier",
    "  Hard   = 5+ étapes ou raisonnement complexe (typique d'un expert)",
    "",
    "⚠ Bonnes pratiques",
    "• Lancez les 3 IA en aveugle (un assistant ne sait pas ce que les 2 autres disent).",
    "• Évitez les prompts qui révèlent l'origine (« ChatGPT, peux-tu... »).",
    "• Pour les calculs : faites le calcul manuellement avant pour avoir la vérité terrain.",
    "• Pour la rédaction : faites lire la réponse à un humain métier (DAF, RH, juriste).",
    "• Pour le multimodal : préparez les fichiers / images en amont (cf. col. Préparation).",
    "",
    "📊 Volumétrie suggérée",
    "Faites au moins 3 prompts par catégorie pour que la moyenne soit représentative.",
    "Idéal : 30 prompts au total, durée du bench ~ 2h.",
    "",
    "🎁 Onglets",
    "  🧮 Calculs métier            : TVA, IS, CAF, marge, indemnités, BFR",
    "  📝 Génération documents      : devis, contrat, fiche poste, RGPD, audit cyber",
    "  📚 Q&R RAG                   : Q&A sur documents internes (cf. seed-data du repo)",
    "  💻 Code & Formules           : Excel, SQL, Python, VBA",
    "  👁️ Vision (image en entrée)  : OCR facture, lecture graphique, tableau scanné",
    "  🎨 Génération d'image        : logo, illustration, infographie (CLOUD only)",
    "  🎤 Voix (TTS / STT)          : dictée, lecture, full hands-free",
    "  💬 Conversation multi-tour   : dialogue, contexte, qualification",
    "  📊 Score final               : récap moyennes par catégorie",
]
for i, line in enumerate(intro, 2):
    ws[f"A{i}"] = line
    if line.startswith(("🎯", "📋", "🏆", "🚦", "⚠", "📊", "🎁")):
        ws[f"A{i}"].font = Font(bold=True, size=12)
ws["A1"].alignment = Alignment(horizontal="center")


# ============================================================================
# ONGLET 2 : CALCULS MÉTIER
# ============================================================================
prompts_calc = [
    {"id": "C01", "level": "Easy", "prompt":
     "Calcule la TVA à 20% sur 1234.56€ HT et donne-moi le détail HT, TVA, TTC.",
     "expected": "HT 1234,56 / TVA 246,91 / TTC 1481,47"},
    {"id": "C02", "level": "Easy", "prompt":
     "Convertis 12 500€ HT en TTC avec une TVA à 5,5%.",
     "expected": "TTC = 13 187,50 €"},
    {"id": "C03", "level": "Medium", "prompt":
     "Une SARL fait 152 800€ de CA HT en services. À partir de quel CA bascule-t-elle "
     "en régime réel normal de TVA ?",
     "expected": "Seuil RSI services 2026 : 247 000€. Donc encore en RSI."},
    {"id": "C04", "level": "Medium", "prompt":
     "Calcule la marge brute, marge nette et taux de marge sur ces ventes : "
     "CA HT 245 600€, achats 134 800€, charges variables 18 200€.",
     "expected": "MB = 110 800 / MN = 92 600 / Tx marge brute = 45,1%"},
    {"id": "C05", "level": "Hard", "prompt":
     "Tu es expert-comptable. Une SARL de 12 salariés a réalisé en 2025 : CA HT 487 320€, "
     "achats 156 800€, salaires bruts 234 500€ (charges patronales 42%), services extérieurs "
     "28 400€, dotations amortissements 18 700€, charges financières 4 200€, produits financiers "
     "1 850€. Calcule pas-à-pas : 1) Résultat exploitation, 2) Résultat financier, 3) RCAI, "
     "4) IS PME (15% jusqu'à 42 500€ puis 25% au-delà), 5) Résultat net, 6) CAF. "
     "Pour chaque étape : formule, calcul, montant, commentaire.",
     "expected": "RE -49 570 / RCAI -51 920 / IS 0 (perte) / Net -51 920 / CAF -33 220"},
    {"id": "C06", "level": "Hard", "prompt":
     "Un commercial vend 18 unités à 1 250€ HT pièce. Le client demande 12% de remise, "
     "paiement à 60 jours. Coût d'achat unitaire : 780€. Frais variables livraison : "
     "95€/unité. L'opération est-elle rentable ? Calcule la marge nette absolue et en %, "
     "et l'impact trésorerie du paiement à 60 jours sur le BFR.",
     "expected": "CA net 19 800 / coût total 15 750 / marge 4 050 (20,5%) / BFR + 19 800 sur 60j"},
    {"id": "C07", "level": "Hard", "prompt":
     "Calcule l'indemnité de licenciement légale d'un salarié non-cadre : 7 ans 3 mois "
     "d'ancienneté, salaire de référence 2 850€ brut/mois. Précise la formule et les barèmes.",
     "expected": "1/4 mois × 7,25 ans × 2850 ≈ 5 165€"},
]
make_prompts_sheet("Calculs métier", "🧮", prompts_calc)


# ============================================================================
# ONGLET 3 : GÉNÉRATION DE DOCUMENT
# ============================================================================
prompts_doc = [
    {"id": "D01", "level": "Easy", "prompt":
     "Rédige un email de relance commercial poli pour un client qui n'a pas répondu à un devis "
     "envoyé il y a 15 jours. Ton professionnel mais chaleureux.",
     "expected": "Email structuré, formule politesse, rappel devis, CTA clair"},
    {"id": "D02", "level": "Medium", "prompt":
     "Rédige un devis pour 3 jours de conseil SEO à 850€/jour HT pour la SARL Bonjour "
     "(SIRET 123 456 789, 14 rue de la Paix Paris 75002), prestations en juin 2026, TVA 20%, "
     "conditions standard (acompte 30%, solde à 30 jours, frais déplacement 0,545€/km).",
     "expected": "HT 2550 / TVA 510 / TTC 3060, mentions légales, conditions"},
    {"id": "D03", "level": "Medium", "prompt":
     "Rédige un contrat de prestation de service simple entre deux SAS françaises pour la "
     "réalisation d'un site web (dev + design, 8 semaines, 24 000€ HT, 30% acompte, "
     "livraison phasée). Inclus : propriété intellectuelle, garantie, résiliation.",
     "expected": "Structure contrat, clauses essentielles, droit français"},
    {"id": "D04", "level": "Medium", "prompt":
     "Rédige une fiche de poste pour un Responsable Commercial PME (4-5 ans XP B2B SaaS, "
     "Île-de-France, 55-65k€ + variable). Inclus missions, profil recherché, KPIs, processus "
     "de recrutement (3 entretiens).",
     "expected": "Fiche structurée, missions, KPIs, processus"},
    {"id": "D05", "level": "Hard", "prompt":
     "Rédige une politique RGPD complète pour un site e-commerce français vendant des "
     "compléments alimentaires (B2C, paiement Stripe, livraison Mondial Relay, mailing Brevo, "
     "pixel Meta). Inclus : finalités, bases légales, durées de conservation, droits "
     "utilisateurs, DPO contact, cookies, transferts hors UE.",
     "expected": "Politique complète, citations RGPD art. 6, 13-22, 32-34, 44-49"},
    {"id": "D06", "level": "Hard", "prompt":
     "Rédige un rapport d'analyse de risque cyber pour une TPE de 15 personnes (compta interne, "
     "clients sensibles, télétravail 50%, cloud Google Workspace, NAS local Synology). "
     "Identifie les 10 risques principaux par ordre de gravité, propose des contre-mesures "
     "concrètes et chiffre le coût annuel d'une politique cyber acceptable.",
     "expected": "10 risques (phishing, ransomware, NAS exposé...) + contre-mesures + budget 5-15k€"},
]
make_prompts_sheet("Génération documents", "📝", prompts_doc)


# ============================================================================
# ONGLET 4 : Q&R RAG
# ============================================================================
prompts_rag = [
    {"id": "R01", "level": "Easy", "prompt":
     "Selon notre procédure congés interne, combien de temps de préavis pour un congé d'1 jour ?",
     "multimodal": "Préalable : importer dans /documents le fichier "
                   "seed-data/01-procedure-conges.md du repo BoxIA.",
     "expected": "24 heures (cf. seed doc procédure congés)"},
    {"id": "R02", "level": "Medium", "prompt":
     "Quel est le taux de TVA pour la restauration consommée sur place selon notre FAQ TVA ?",
     "multimodal": "Préalable : importer seed-data/03-faq-tva.md.",
     "expected": "10% (cf. FAQ TVA seed)"},
    {"id": "R03", "level": "Medium", "prompt":
     "Quels sont les seuils 2026 du régime simplifié de TVA pour une activité mixte "
     "(biens + services) ?",
     "multimodal": "FAQ TVA seed.",
     "expected": "85 800€ biens / 34 400€ services / 818 000€ et 247 000€ pour bascule réel normal"},
    {"id": "R04", "level": "Hard", "prompt":
     "À partir de notre modèle de devis interne, génère un devis pour 5 jours de prestation "
     "de conseil en stratégie à 1 200€/jour, client SARL Acme, TVA 20%. Reprends nos "
     "mentions légales du modèle.",
     "multimodal": "Importer seed-data/02-modele-devis.md.",
     "expected": "Devis structuré qui reprend les mentions de la KB"},
    {"id": "R05", "level": "Hard", "prompt":
     "Tu es directeur RH. Compare notre procédure congés avec la convention collective Syntec. "
     "Quels sont les écarts à corriger ?",
     "multimodal": "Avoir importé la procédure interne ET la convention Syntec (à charger).",
     "expected": "Identification des divergences délais préavis / fractionnement / RTT"},
]
make_prompts_sheet("Q&R RAG (documents internes)", "📚", prompts_rag)


# ============================================================================
# ONGLET 5 : CODE & FORMULES
# ============================================================================
prompts_code = [
    {"id": "K01", "level": "Easy", "prompt":
     "Donne-moi la formule Excel pour additionner les cellules B2:B100 si la cellule "
     "correspondante en A vaut « oui ».",
     "expected": "=SOMME.SI(A2:A100;\"oui\";B2:B100)"},
    {"id": "K02", "level": "Medium", "prompt":
     "Écris une requête SQL pour trouver les 10 clients qui ont le plus dépensé en 2025, "
     "table `commandes(client_id, date, montant_ht)`, jointure avec `clients(id, nom, email)`.",
     "expected": "SELECT c.nom, SUM(o.montant_ht)... GROUP BY... ORDER BY... DESC LIMIT 10"},
    {"id": "K03", "level": "Medium", "prompt":
     "Écris un script Python qui lit un CSV de factures (numero, date, client, montant_ht, "
     "tva_taux, statut) et génère un rapport mensuel des CA HT et TVA collectée groupé "
     "par client. Sortie : DataFrame pandas + export Excel.",
     "expected": "Script pandas avec read_csv, groupby, to_excel, gestion erreurs"},
    {"id": "K04", "level": "Hard", "prompt":
     "Rédige une macro VBA Excel qui automatise l'envoi d'un mail Outlook avec en pièce "
     "jointe le PDF généré depuis l'onglet « Devis ». Destinataire en B5, objet en B6, "
     "corps en B7. Gère les erreurs (Outlook fermé, PDF non généré).",
     "expected": "Sub VBA avec ExportAsFixedFormat, CreateItem, On Error..."},
]
make_prompts_sheet("Code & Formules", "💻", prompts_code)


# ============================================================================
# ONGLET 6 : VISION
# ============================================================================
prompts_vision = [
    {"id": "V01", "level": "Easy", "prompt":
     "Décris ce que tu vois dans cette image. Sois précis sur les éléments visibles "
     "(objets, personnes, contexte, couleurs).",
     "multimodal": "Joindre une photo de bureau de TPE (post-it, écran, café). "
                   "BoxIA = Assistant général uniquement (qwen2.5vl). "
                   "ChatGPT = trombone. Gemini = +.",
     "expected": "Description structurée et précise, identification des objets clés"},
    {"id": "V02", "level": "Medium", "prompt":
     "Voici une facture fournisseur. Extrais en tableau Markdown : numéro, date, "
     "fournisseur, total HT, TVA, total TTC, échéance.",
     "multimodal": "Joindre une vraie facture PDF/PNG (idéalement scannée pour tester l'OCR). "
                   "BoxIA : qwen2.5vl uniquement.",
     "expected": "Tableau structuré, valeurs lues correctement"},
    {"id": "V03", "level": "Medium", "prompt":
     "Voici un screenshot d'un graphique financier. Décris-le : axes, tendances, "
     "points remarquables, conclusion business.",
     "multimodal": "Joindre un graphique (CA mensuel, courbe acquisition).",
     "expected": "Lecture correcte du graphique + interprétation"},
    {"id": "V04", "level": "Hard", "prompt":
     "Voici la photo d'un tableau Excel (capture d'écran) avec un échéancier de paiements. "
     "Reconstitue ce tableau en Markdown et donne le total à payer ce mois.",
     "multimodal": "Joindre un screenshot Excel ~10 lignes d'échéances. "
                   "Test critique pour la fonction « OCR vers tableur » en démo.",
     "expected": "Tableau reconstitué fidèle, total correct"},
    {"id": "V05", "level": "Hard", "prompt":
     "Sur cette photo d'une affiche/flyer concurrent, identifie : 1) la promesse principale, "
     "2) le call-to-action, 3) le pricing affiché, 4) les éléments visuels qui attirent "
     "l'œil. Puis propose 3 idées pour différencier notre offre.",
     "multimodal": "Joindre la photo d'une affiche commerciale.",
     "expected": "Analyse marketing structurée + propositions différenciation"},
]
make_prompts_sheet("Vision (image en entrée)", "👁️", prompts_vision)


# ============================================================================
# ONGLET 7 : GÉNÉRATION D'IMAGE (cloud only)
# ============================================================================
prompts_imggen = [
    {"id": "I01", "level": "Easy", "prompt":
     "Crée un logo simple, plat, dans des tons bleus pour une entreprise nommée "
     "« AquaTPE » (plomberie pour TPE).",
     "multimodal": "BoxIA : ❌ pas dispo nativement. ChatGPT : commande directe. "
                   "Gemini : « Créer une image » dans la barre.",
     "expected": "BoxIA = N/A (note 0). Cloud : visuel acceptable on-brief"},
    {"id": "I02", "level": "Medium", "prompt":
     "Génère une illustration vectorielle pour une newsletter SaaS RH en flat-design, thème "
     "« équipe heureuse en télétravail », tons pastels, style Slack/Notion.",
     "multimodal": "Cloud uniquement.",
     "expected": "Cloud : illustration cohérente, on-brief"},
    {"id": "I03", "level": "Hard", "prompt":
     "Crée une infographie présentant les 4 étapes d'un parcours d'achat e-commerce "
     "(Découverte → Considération → Achat → Fidélisation) avec icônes minimalistes "
     "et palette monochrome verte.",
     "multimodal": "Cloud uniquement. Test cohérence + lisibilité texte généré.",
     "expected": "Cloud : infographie lisible, structurée, 4 étapes claires"},
]
make_prompts_sheet("Génération d'image", "🎨", prompts_imggen)


# ============================================================================
# ONGLET 8 : VOIX
# ============================================================================
prompts_voice = [
    {"id": "VX01", "level": "Easy", "prompt":
     "Dicte cette phrase à voix normale : « Bonjour, je voudrais un devis pour 5 jours "
     "de conseil en management à compter de juin. »",
     "multimodal": "BoxIA : cliquer le micro et dicter. ChatGPT : icône micro. "
                   "Gemini : pareil. Mesurer précision transcription FR.",
     "expected": "Transcription fidèle (idéalement 100% sans accents perdus)"},
    {"id": "VX02", "level": "Medium", "prompt":
     "Lis-moi à haute voix la réponse précédente.",
     "multimodal": "BoxIA : bouton volume sur la réponse. ChatGPT : Read aloud. "
                   "Gemini : variable selon device.",
     "expected": "TTS lisible, voix française naturelle"},
    {"id": "VX03", "level": "Hard", "prompt":
     "Dicte un texte de 1 minute (compte-rendu de réunion fictif), puis demande à l'IA "
     "de résumer en 5 points et lis-moi le résumé.",
     "multimodal": "Test full hands-free : dictée + résumé + lecture vocale.",
     "expected": "Pipeline complet sans clavier, résumé pertinent, lecture fluide"},
]
make_prompts_sheet("Voix (TTS - STT)", "🎤", prompts_voice)


# ============================================================================
# ONGLET 9 : CONVERSATION MULTI-TOUR
# ============================================================================
prompts_conv = [
    {"id": "M01", "level": "Medium", "prompt":
     "Conversation en 5 tours pour qualifier le besoin d'un prospect TPE :\n"
     "1) « Bonjour, je veux digitaliser mon entreprise. »\n"
     "2) « On est 8 personnes, secteur BTP. »\n"
     "3) « Pas de logiciel actuellement, juste Excel et papier. »\n"
     "4) « Budget environ 500€/mois max. »\n"
     "5) « Besoin urgent : facturation et suivi chantier. »",
     "multimodal": "Envoyer chaque message un par un. Voir si l'IA garde le contexte.",
     "expected": "Réponse finale qui synthétise les 5 réponses, qualifie le besoin, "
                 "propose un plan d'action chiffré"},
    {"id": "M02", "level": "Hard", "prompt":
     "Tu es expert juridique. Je vais te poser plusieurs questions sur la rupture "
     "conventionnelle d'un de mes salariés. Tu réponds en posant les bonnes questions "
     "de clarification AVANT de donner ton avis. Première question : « Mon salarié "
     "veut partir, peut-il imposer une RC ? »",
     "multimodal": "Test de la capacité à dialoguer (pas juste répondre).",
     "expected": "L'IA pose des questions au lieu de juste répondre. Top : "
                 "ancienneté ? motif sous-jacent ? précédent désaccord ?"},
]
make_prompts_sheet("Conversation multi-tour", "💬", prompts_conv)


# ============================================================================
# ONGLET 10 : SCORE FINAL
# ============================================================================
ws_score = wb.create_sheet("📊 Score final")
ws_score["A1"] = "Synthèse des notes (à remplir au fur et à mesure du bench)"
ws_score["A1"].font = Font(bold=True, size=14)
ws_score.column_dimensions["A"].width = 30
ws_score.column_dimensions["B"].width = 14
ws_score.column_dimensions["C"].width = 14
ws_score.column_dimensions["D"].width = 14
ws_score.column_dimensions["E"].width = 50

headers_score = ["Catégorie", "BoxIA moy/5", "ChatGPT moy/5", "Gemini moy/5", "Verdict"]
for col, h in enumerate(headers_score, 1):
    c = ws_score.cell(row=3, column=col, value=h)
    c.font = HEADER
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = THIN

cats = [
    "🧮 Calculs métier",
    "📝 Génération documents",
    "📚 Q&R RAG",
    "💻 Code & Formules",
    "👁️ Vision (image en entrée)",
    "🎨 Génération d'image",
    "🎤 Voix (TTS - STT)",
    "💬 Conversation multi-tour",
]
for i, label in enumerate(cats, 4):
    ws_score.cell(row=i, column=1, value=label)
    for col in range(2, 6):
        ws_score.cell(row=i, column=col, value="").border = THIN
    ws_score.cell(row=i, column=1).border = THIN

# Conseils finaux
hints = [
    "",
    "📌 Comment lire les scores",
    "  • Score BoxIA ≥ 3,5 sur une catégorie → vous pouvez utiliser BoxIA en confiance.",
    "  • Score BoxIA < 3 → préférez ChatGPT/Gemini en complément, OU upgradez le modèle",
    "    (qwen2.5:14b ou 32b si VRAM dispo, ou Mistral Large via API).",
    "  • Différence > 1,5 point entre BoxIA et le cloud → fix prioritaire (pre-prompt + RAG).",
    "",
    "💡 La différenciation BoxIA reste : privacy + coût récurrent zéro + hors-ligne.",
    "   Pour les 80% de cas usage où le score ≥ 3,5, la box gagne malgré une qualité brute",
    "   légèrement inférieure au cloud sur les tâches complexes.",
    "",
    "🔧 Quand BoxIA est faible sur une catégorie, options par ordre de coût :",
    "   1. Renforcer le pre-prompt de l'agent (gratuit)",
    "   2. Importer plus de contexte dans la KB / RAG (gratuit)",
    "   3. Switch vers un modèle plus gros : 7b → 14b → 32b (gratuit, juste VRAM)",
    "   4. Acheter une 2e GPU pour modèles ≥ 70b (~2 000 €)",
]
start = 4 + len(cats) + 2
for i, h in enumerate(hints, start):
    c = ws_score.cell(row=i, column=1, value=h)
    if h.startswith(("📌", "💡", "🔧")):
        c.font = Font(bold=True, size=12)


wb.save(OUT)
print(f"✓ Saved: {OUT}")
import os
print(f"  Size: {os.path.getsize(OUT) / 1024:.1f} KB")
print(f"  Sheets: {wb.sheetnames}")
